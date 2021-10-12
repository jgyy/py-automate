"""
Sends emails based on their status in spreadsheet.
"""
from sys import argv
from smtplib import SMTP
from os.path import dirname, join
from openpyxl import load_workbook


def wrapper():
    """
    Open the spreadsheet and get the latest dues status.
    Check each member's payment status
    Log in to email account.
    Send out reminder emails.
    """
    path = dirname(__file__)
    wbook = load_workbook(join(path, "dues_records.xlsx"))
    sheet = wbook["Sheet1"]
    last_col = sheet.max_column
    latest_month = sheet.cell(row=1, column=last_col).value
    unpaid_members = {}
    for row in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=row, column=last_col).value
        if payment != "paid":
            name = sheet.cell(row=row, column=1).value
            email = sheet.cell(row=row, column=2).value
            unpaid_members[name] = email
    smtp_obj = SMTP("smtp.gmail.com", 587)
    smtp_obj.ehlo()
    smtp_obj.starttls()
    smtp_obj.login("my_email_address@gmail.com", argv[1])
    for name, email in unpaid_members.items():
        body = (
            f"Subject: {latest_month} dues unpaid.\nDear {name},"
            + "\nRecords show that you have not paid dues for {latest_month}. "
            + "Please make this payment as soon as possible. Thank you!"
        )
        print("Sending email to %s..." % email)
        sendmail_status = smtp_obj.sendmail("my_email_address@gmail.com", email, body)
        if sendmail_status != {}:
            print(f"There was a problem sending email to {email}: {sendmail_status}")
    smtp_obj.quit()


if __name__ == "__main__":
    wrapper()
