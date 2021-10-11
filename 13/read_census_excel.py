"""
Tabulates population and number of census tracts for each county.
"""
from os.path import dirname
from pprint import pformat
from openpyxl import load_workbook


def wrapper():
    """
    Fill in countyData with each county's population and tracts.
    Each row in the spreadsheet has data for one census tract.
    Make sure the key for this state exists.
    Make sure the key for this county in this state exists.
    Each row represents one census tract, so increment by one.
    Increase the county pop by the pop in this census tract.
    Open a new text file and write the contents of countyData to it.
    """
    path = dirname(__file__) + "\\"
    print("Opening workbook...")
    wbook = load_workbook(f"{path}censuspopdata.xlsx")
    sheet = wbook["Population by Census Tract"]
    county_data = {}
    print("Reading rows...")
    for row in range(2, sheet.max_row + 1):
        state = sheet["B" + str(row)].value
        county = sheet["C" + str(row)].value
        pop = sheet["D" + str(row)].value
        county_data.setdefault(state, {})
        county_data[state].setdefault(county, {"tracts": 0, "pop": 0})
        county_data[state][county]["tracts"] += 1
        county_data[state][county]["pop"] += int(pop)
    print("Writing results...")
    with open(f"{path}census2010.py", "w", encoding="utf-8") as result_file:
        result_file.write("all_data = " + pformat(county_data))
    print("Done.")


if __name__ == "__main__":
    wrapper()
