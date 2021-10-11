"""
Corrects costs in produce sales spreadsheet.
"""
from os.path import dirname
from openpyxl import load_workbook


def wrapper():
    """
    The produce types and their updated prices
    Loop through the rows and update the prices
    skip the first row
    """
    path = dirname(__file__) + "\\"
    wbook = load_workbook(f"{path}produce_sales.xlsx")
    sheet = wbook["Sheet"]
    price_updates = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}
    for row_num in range(2, sheet.max_row):
        produce_name = sheet.cell(row=row_num, column=1).value
        if produce_name in price_updates:
            sheet.cell(row=row_num, column=2).value = price_updates[produce_name]
    wbook.save(f"{path}updated_produce_sales.xlsx")


if __name__ == "__main__":
    wrapper()
